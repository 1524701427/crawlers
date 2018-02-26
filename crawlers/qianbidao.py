# -*- coding: utf-8 -*-

"""爬取铅笔道项目的爬虫。"""

from __future__ import print_function


import time
import datetime

import openpyxl

from const import (
    HTTP_OK,
)
from config import (
    QIANBIDAO_USER,
    QIANBIDAO_PASSWORD,
    RECEIPTS,
)
from util.http import HttpClient
from util.mail import mail_multipart


class QianBiDaoCrawler(object):

    QianBiDaoAppHeaders = {
        'User-Agent': 'PencilNews/1.3.0 (iPhone; iOS 11.2.5; Scale/2.00)',
        'Accept-Encoding': 'br, gzip, deflate',
        'Accept-Language': 'zh-Hans-CN;q=1, en-CN;q=0.9',
        'Authority': 'api.pencilnews.cn',
    }

    def __init__(self, username, password):
        self.httpclient = HttpClient(default_headers=self.QianBiDaoAppHeaders)
        self.username = username
        self.password = password
        self.xlsx_name = ""
        self.token = ""

    def send_mail(self):
        """发送邮件"""
        mail = dict()
        mail['to'] = RECEIPTS
        mail['subject'] = self.xlsx_name
        mail['attachment'] = [u'%s.xlsx' % self.xlsx_name]
        mail_multipart(mail)

    def login(self):
        """登录铅笔道"""
        data = dict(username=self.username, password=self.password)
        resp = self.httpclient.post(
            'https://api.pencilnews.cn/user/login',
            data=data,
            verify=False)
        if resp and resp.status_code == HTTP_OK:
            ret = resp.json()
            if ret and ret['message'] == 'SUCCESS':
                self.token = ret['data']['user']['token']
        return self.token

    def get_item_detail(self, item_id):
        """获取项目详情。"""
        url = 'https://api.pencilnews.cn/pay-project/detail?id=%s' % item_id
        print('>>>>>>', url)
        resp = self.httpclient.get(url, headers=dict(token=self.token))
        if resp and resp.status_code == HTTP_OK:
            data = resp.json()
            if data['message'] == 'SUCCESS':
                return data['data']
        return {}

    def run(self):
        """爬虫运行"""
        page = 1
        url_tpl = 'https://api.pencilnews.cn/pay-project/list?page=%(page)d'
        projects = list()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.titile = u'铅笔道项目列表'
        headers = [
            u'项目名称', u'区域', u'细分领域', u'官网',
            u'项目简介', u'融资轮次', u'融资情况', u'项目来源']
        for idx, header in enumerate(headers, 1):
            sheet.cell(row=1, column=idx, value=header)
        while True:
            url = url_tpl % dict(page=page)
            print('>>>', url)
            resp = self.httpclient.get(url, headers=dict(token=self.token))
            if resp and resp.status_code == HTTP_OK:
                data = resp.json()
                for item in data['data']['items']:
                    project = dict()
                    project['id'] = item['id']
                    project['name'] = item['name']
                    project['description'] = item['description']
                    project['web'] = ''
                    project['industry'] = item['industry'][0]
                    project['finance'] = item['finance_round'] or u'尚未获投'
                    time.sleep(1)
                    detail_info = self.get_item_detail(project['id'])
                    detail_info = detail_info['content']
                    project['region'] = detail_info['project']['region_name']
                    rounds = list()
                    for _round in detail_info.get('rounds', []):
                        rounds.append('%s %s %s %s\n' % (
                            _round['annouced_time'],
                            _round['stage_name'],
                            _round['money_raised'],
                            _round['investor']))
                    project['finance_round'] = '\n'.join(rounds)
                    projects.append(project)
            time.sleep(1)
            page += 1
            if page > 1:
                break
        for idx, project in enumerate(projects, 2):
            sheet.cell(row=idx, column=1, value=project['name'])
            sheet.cell(row=idx, column=2, value=project['region'])
            sheet.cell(row=idx, column=3, value=project['industry'])
            sheet.cell(row=idx, column=4, value=project['web'])
            sheet.cell(row=idx, column=5, value=project['description'])
            sheet.cell(row=idx, column=6, value=project['finance'])
            sheet.cell(row=idx, column=7, value=project['finance_round'])
            sheet.cell(row=idx, column=8, value=u'铅笔道')
        self.xlsx_name = u'%s 铅笔道项目' % datetime.date.today()
        workbook.save(u'%s.xlsx' % self.xlsx_name)
        self.send_mail()


if __name__ == '__main__':
    cralwer = QianBiDaoCrawler(QIANBIDAO_USER, QIANBIDAO_PASSWORD)
    cralwer.login()
    cralwer.run()
