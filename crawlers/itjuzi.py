# -*- coding: utf-8 -*-
# flake8: noqa

'''
爬取it桔子项目的爬虫'''

import sys
import time
import random
from datetime import date

import requests
import openpyxl
from openpyxl.styles import Font, Border, Side, NamedStyle, Alignment
from bs4 import BeautifulSoup

import config
from mail import mail_multipart

UAs = [
    'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50',
    'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0;',
    'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0)',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1',
    'Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11',
]

headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Encoding': 'gzip, deflate, sdch, br',
    'User-Agent': random.choice(UAs),
}


class HttpClient(object):

    def __init__(self, default_headers=None, tries=3, try_internal=0.5):
        self.tries = tries
        self.try_internal = 3
        self.s = requests.Session()
        if default_headers is not None:
            self.s.headers.update(default_headers)

    def get(self, url):
        for i in range(1, self.tries+1):
            try:
                self.s.headers.update({'User-Agent': random.choice(UAs)})
                resp = self.s.get(url)
                break
            except:
                time.sleep(self.try_internal*i)
                continue
        else:
            raise RuntimeError("bad network...")
        return resp

    def login(self, user, password):
        url = 'https://www.itjuzi.com/user/login?redirect=&flag='
        content_type = 'application/x-www-form-urlencoded'
        data = {
            'identity': user,
            'password': password,
            'remember': 1,
            'page': None,
            'url': None,
        }
        resp = self.s.post(url, data=data, headers={'Content-Type': content_type})
        if resp is not None:
            for c in resp.cookies:
                print('>', c.name, '=', c.value)


def export(projects):
    subject = u'%s 日it桔子项目汇总' % date.today()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = u'IT桔子'
    # 设置字体
    font = Font(name=u'楷体', size=10)
    # 设置对齐方式
    alignment = Alignment(vertical='center', wrap_text=True)
    # 自定义样式
    stdstyle = NamedStyle(name=u'标准样式')
    stdstyle.font = font
    stdstyle.alignment = alignment
    xls_headers = [
        u'项目名称', u'地址',  u'细分领域', u'项目连接', u'项目简介', u'轮次', u'融资情况'
    ]
    for i, header in enumerate(xls_headers, 1):
        sheet.cell(row=1, column=i, value=header)

    row = 1
    for project in projects:
        row += 1
        cell = sheet.cell(row=row, column=1, value=project['name'])
        cell.style = stdstyle
        cell = sheet.cell(row=row, column=2, value=project['location'])
        cell.style = stdstyle
        cell = sheet.cell(row=row, column=3, value=project['industry'])
        cell.style = stdstyle
        cell = sheet.cell(row=row, column=4, value=project['web'])
        cell.style = 'Hyperlink'  # builtin样式，超链接样式
        cell.alignment = alignment
        cell = sheet.cell(row=row, column=5, value=project['abstract'])
        cell.font = font
        cell.alignment = Alignment(wrap_text=True)
        financing_round = project['investing']
        cell = sheet.cell(row=row, column=6, value=financing_round)
        cell.style = stdstyle
        cell = sheet.cell(row=row, column=7, value=project['investing'])
        last_round = None
        if project['financings']:
            last_round = project['financings'][0]
        if financing_round not in [u'尚未获投', u'获投状态：不明确'] and last_round:
            cell = sheet.cell(
                row=row, column=7, value= '%s %s %s %s' % (
                    last_round['date'],
                    last_round['round'],
                    last_round['fee'],
                    ','.join(last_round['investors'])))
        cell.style = stdstyle
    # 自动调整列宽
    dims = dict()
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = \
                    max((dims.get(cell.column, 0), len(cell.value)))
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value + 2
    # 设置列宽
    sheet.column_dimensions['E'].width = 60
    workbook.save(u'%s.xlsx' % subject)

    email = dict()
    email['to'] = config.RECEIPTS
    email['subject'] = subject
    email['attachment'] = [u'%s.xlsx' % subject]
    mail_multipart(email)


def crawler(user, password):
    '''it桔子爬虫'''
    client = HttpClient(headers)
    # client.login(user, password)
    init_page = page = 0
    delimiters = '>'*10
    url_tpl = (
        'https://www.itjuzi.com/company?sortby=inputtime&page=%(page)d')
    quit = False
    projects = []
    while quit is False:
        page += 1
        url = url_tpl % dict(page=page)
        print(delimiters, url)
        resp = client.get(url)
        if resp is not None:
            soup = BeautifulSoup(resp.text, 'lxml')
            print(soup)

            # tag_ul = soup.select('ul[class="list-main-icnset company-list-ul"]')[0]
            tag_ul = soup.select('ul[class="list-main-icnset company-list-ul"]')[0]
            for idx, tag_li in enumerate(tag_ul.find_all('li')):
                try:
                    tag_is = [x for x in tag_li.find_all('i')]
                    tag_spans = [x for x in tag_li.find_all('span')]

                    project = dict()
                    project['url'] = tag_is[0].a['href']
                    print(project['url'])

                    project['location'] = ''
                    _is = tag_li.select('i[class="cell place"]')
                    if _is:
                        project['location'] = _is[0].string.strip()

                    project['name'] = ''
                    project['industry'] = ''

                    time.sleep(15)
                    detail_url = project['url']
                    detail_resp = client.get(detail_url)

                    detail_soup = BeautifulSoup(detail_resp.text, 'lxml')
                    #  locations = detail_soup.select('span[class="loca c-gray-aset"]')
                    #  if locations:
                    #      try:
                    #          project['location'] = locations[0].a.string.strip()
                    #      except:
                    #          pass

                    industry = detail_soup.select('a[class="one-level-tag"]')
                    if industry:
                        try:
                            project['industry'] = industry[0].string.strip()
                        except:
                            pass

                    name = detail_soup.select('h1[class="seo-important-title"]')
                    if name:
                        try:
                            project['name'] = tag_li.div.div.a.span.string.strip()
                        except:
                            pass

                    # 项目状态
                    project['investing'] = ''
                    title_blocks = detail_soup.select('span[class="t-small c-green"]')
                    if title_blocks:
                        # 去除掉两边的"(", ")"
                        project['investing'] = title_blocks[0].string.strip()[1:-1]

                    div_link_line = detail_soup.select('div[class="link-line"]')[0]
                    project['web'] = ''
                    web_links = div_link_line.select('a[target="_blank"]')
                    for web_link in web_links:
                        if web_link['href']:
                            project['web'] = web_link['href'].strip()
                            break

                    project['abstract'] = detail_soup.find(attrs={"name": "Description"})['content']
                    project['abstract'] = project['abstract'].strip()
                    if project['abstract'][-1] not in (u'。', u'.'):
                        project['abstract'] = project['abstract'] + u'。'

                    financings = []
                    tables = detail_soup.select('table[class="list-round-v2"]')
                    if tables:
                        table = tables[0]
                        for tr in table.find_all('tr'):
                            financing = dict()
                            tds = [x for x in tr.find_all("td")]
                            financing['date'] = tds[0].span.string
                            financing['round'] = tds[1].span.string
                            financing['fee'] = tds[2].span.string
                            financing['investors'] = [x for x in tds[3].strings if x != '\n']
                            financings.append(financing)
                    project['financings'] = financings
                    # print(project)
                    projects.append(project)
                except:
                    pass
        time.sleep(15)
        # itjuzi最多可以爬50页数据
        if page - init_page >= 8:
            break
    return projects


if __name__ == '__main__':
    projects = crawler(config.ITJUZI_USER, config.ITJUZI_PASSWORD)
    export(projects)
