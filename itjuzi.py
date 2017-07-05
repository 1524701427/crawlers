# -*- coding: utf-8 -*-
# flake8: noqa

'''
爬取it桔子项目的爬虫'''

from __future__ import print_function

import sys
import time
from datetime import date

import requests
import openpyxl
from openpyxl.styles import Font, Border, Side, NamedStyle, Alignment
from bs4 import BeautifulSoup

import config
from mail import mail_multipart

headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Encoding': 'gzip, deflate, sdch, br',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36',
}

DEBUG = True


class HttpClient(object):

    def __init__(self, default_headers=None, tries=3, try_internal=0.5):
        assert 1<= tries <= 5
        self.tries = tries
        self.try_internal = 3
        self.s = requests.Session()
        if headers is not None:
            self.s.headers.update(headers)

    def get(self, url):
        for i in range(1, self.tries+1):
            try:
                resp = self.s.get(url)
                break
            except:
                time.sleep(self.try_internal*i)
                continue
        else:
            raise RuntimeError("bad network...")
        return resp

    def login(self, user, password):
        # 建议抽象化为基类，以后使用，增加验证码识别模块
        url = 'https://www.itjuzi.com/user/login?redirect=&flag='
        content_type = 'application/x-www-form-urlencoded'
        data = {
            'identity': user,
            'password': password,
            'remember': 1,
            'page': None,
            'url': None,
        }
        resp = self.s.post(url, data=data, headers={'Content-Type': content_type})
        print(resp)
        if resp is not None:
            for c in resp.cookies:
                print('>', c.name, '=', c.value)


def export(projects):
    subject = u'%s 日it桔子项目汇总' % date.today() 
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = u'IT桔子'
    # 设置字体
    font = Font(name=u'楷体', size=10)
    # 设置对齐方式
    alignment = Alignment(wrap_text=True)
    xls_headers = [
        u'项目名称', u'地址',  u'细分领域', u'项目连接', u'项目简介', u'融资情况'
    ]
    for i, header in enumerate(xls_headers, 1):
        sheet.cell(row=1, column=i, value=header)

    row = 1
    for project in projects:
        row += 1
        cell = sheet.cell(row=row, column=1, value=project['name'])
        cell.font = font
        cell = sheet.cell(row=row, column=2, value=project['location'])
        cell.font = font
        cell = sheet.cell(row=row, column=3, value=project['industry'])
        cell.font = font
        cell = sheet.cell(row=row, column=4, value=project['web'])
        cell.style = 'Hyperlink'  # builtin样式，超链接样式
        cell = sheet.cell(row=row, column=5, value=project['abstract'])
        cell.width = 60
        cell.font = font
        cell.alignment = alignment
    workbook.save(u'%s.xlsx' % subject)

    email = dict()
    email['to'] = ['wanglanwei@weiche.cn']
    email['subject'] = subject
    email['attachment'] = [u'%s.xlsx' % subject]
    mail_multipart(email)


def get_last_id():
    '''获取上次爬取的截止id'''
    try:
        with open('last_id', 'rt') as f:
            last_id = f.read()
            last_id = int(last_id.strip())
    except IOError:
        last_id = 67960
        with open('last_id', 'wt') as f:
            f.write(str(last_id))
        assert last_id > 0
    finally:
        return last_id


def set_last_id(last_id):
    '''设置上次爬去的截止id'''
    with open('last_id', 'wt') as f:
        f.write('%d\n' % last_id)
        return last_id


def crawler(user, password):
    '''it桔子爬虫'''
    client = HttpClient(headers)
    client.login(user, password)
    # last_id = get_last_id()
    init_page = page = 20
    delimiters = '>'*10
    url_tpl = (
        'http://www.itjuzi.com/company?sortby=inputtime&page=%(page)d')
    quit = False
    projects = []
    while quit is False:
        page += 1
        url = url_tpl % dict(page=page)
        print(delimiters, url)
        resp = client.get(url)
        if resp is not None:
            soup = BeautifulSoup(resp.text, 'lxml')
            print(soup)

            tag_ul = soup.select('ul[class="list-main-icnset list-main-com"]')[0]
            for idx, tag_li in enumerate(tag_ul.find_all('li')):
                tag_is = [x for x in tag_li.find_all('i')]
                tag_spans = [x for x in tag_li.find_all('span')]

                project = dict()
                project['url'] = tag_is[0].a['href']
                print(project['url'])

                project_id = int(project['url'].split('/')[-1])
                project['id'] = project_id

                # 运行到截止id退出
                if project_id <= last_id:
                    quit = True
                    break

                project['name'] = tag_li.p.a.string
                project['industry'] = tag_spans[2].a.string
                
                time.sleep(5)
                detail_url = project['url']
                detail_resp = client.get(detail_url)

                detail_soup = BeautifulSoup(detail_resp.text, 'lxml')
                project['location'] = ''
                locations = detail_soup.select('span[class="loca c-gray-aset"]')
                if locations:
                    project['location'] = locations[0].a.string.strip()

                div_link_line = detail_soup.select('div[class="link-line"]')[0]
                project['web'] = ''
                web_links = div_link_line.select('a[target="_blank"]')
                for web_link in web_links:
                    if web_link['href']:
                        project['web'] = web_link['href'].strip()
                        break

                project['abstract'] = detail_soup.find(attrs={"name": "Description"})['content']

                financings = []
                tables = detail_soup.select('table[class="list-round-v2"]')
                if tables:
                    table = tables[0]
                    for tr in table.find_all('tr'):
                        financing = dict()
                        tds = [x for x in tr.find_all("td")]
                        financing['date'] = tds[0].span.string
                        financing['round'] = tds[1].span.string
                        financing['fee'] = tds[2].span.string
                        financing['investors'] = [x for x in tds[3].strings if x != '\n']
                        financings.append(financing)
                project['financings'] = financings
                projects.append(project)
        time.sleep(5)
        if page - init_page >= 10:
            break
    last_id = projects[0]['id']
    set_last_id(last_id)
    return projects


if __name__ == '__main__':
    projects = crawler(config.ITJUZI_USER, config.ITJUZI_PASSWORD)
    export(projects)
